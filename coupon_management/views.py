import re
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from coupon_management.serializers import couponStockSerializers
from lxml.etree import HTML
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate

from client_management.models import *
from competitor_analysis.forms import CompetitorAnalysisFilterForm
from master.functions import generate_form_errors
from product.models import Staff_Orders_details
from .models import *
from .forms import  *
from accounts.models import CustomUser, Customers
from master.models import EmirateMaster, BranchMaster, RouteMaster
import json
from django.core.serializers import serialize
from django.views import View
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from accounts.views import log_activity
from django.db.models import Count, Q


# Create your views here.
def get_next_coupon_bookno(request):
    coupon_type = request.GET.get("coupon_type")
    next_coupon_bookno = ""
    next_leaf_no = ""
    end_leaf_no = ""
    next_free_leaf_no = ""
    end_free_leaf_no = ""
    
    last_coupon = NewCoupon.objects.filter(coupon_type__pk=coupon_type)
    if last_coupon.exists():
        last_coupon = last_coupon.latest("created_date")
        # last_coupon_bookno = last_coupon.book_num
        next_coupon_bookno = int(last_coupon.book_num) + 1

        # Split the alphanumeric book number into alphabetic and numeric parts
        # match = re.match(r"([a-zA-Z]+)(\d+)", last_coupon_bookno)
        # if not match:
        #     raise ValueError("Invalid book number format")

        # alphabetic_part, numeric_part = match.groups()
        # next_numeric_part = int(numeric_part) + 1
        # next_coupon_bookno = f"{alphabetic_part}{next_numeric_part}"
        
        if (leaflet:=CouponLeaflet.objects.filter(coupon=last_coupon)).exists():
            last_leaf_number = leaflet.latest("created_date").leaflet_name
            
            # Split the alphanumeric leaflet number into alphabetic and numeric parts
            # match = re.match(r"([a-zA-Z]+)(\d+)", last_leaf_name)
            # if not match:
            #     raise ValueError("Invalid leaf number format")

            # leaf_alphabetic_part, leaf_name_part = match.groups()
            next_leaf_no = int(last_leaf_number) + 1
            end_leaf_no = next_leaf_no + int(last_coupon.valuable_leaflets) - 1
            
            # next_leaf_no = f"{leaf_alphabetic_part}{leaf_next_numeric_part}"
            # end_leaf_no = f"{leaf_alphabetic_part}{leaf_last_numeric_part}"
            
            if (freeleafs:=FreeLeaflet.objects.filter(coupon=last_coupon)).exists():
                last_free_leaf_number = freeleafs.latest("created_date").leaflet_name
                
                # Split the alphanumeric leaflet number into alphabetic and numeric parts
                # match = re.match(r"([a-zA-Z]+)(\d+)", last_free_leaf_name)
                # if not match:
                #     raise ValueError("Invalid leaf number format")

                # free_leaf_alphabetic_part, free_leaf_name_part = match.groups()
                free_leaf_next_number = int(last_free_leaf_number) + 1
                if int(last_coupon.free_leaflets) > 1:
                    free_leaf_last_number = free_leaf_next_number + int(last_coupon.free_leaflets) - 1
                else:
                    free_leaf_last_number = free_leaf_next_number
                
                next_free_leaf_no = f"{free_leaf_next_number}"
                end_free_leaf_no = f"{free_leaf_last_number}"

    data = {
        'next_coupon_bookno': next_coupon_bookno,
        "next_leaf_no": next_leaf_no,
        "end_leaf_no": end_leaf_no,
        "next_free_leaf_no": next_free_leaf_no,
        "end_free_leaf_no": end_free_leaf_no,
        }
    return JsonResponse(data, safe=False)


def get_coupon_bookno(request):
    request_id = request.GET.get("request_id")
    
    if (instances:=Staff_Orders_details.objects.filter(pk=request_id)).exists():
        instance = instances.first()
        stock_instances = CouponStock.objects.filter(couponbook__coupon_type__coupon_type_name=instance.product_id.product_name,coupon_stock="company")
        serialized = couponStockSerializers(stock_instances, many=True)
        
        status_code = 200
        response_data = {
            "status": "true",
            "data": serialized.data,
        }
    else:
        status_code = 404
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

def couponType(request):
    all_couponType = CouponType.objects.all()
    context = {'all_couponType': all_couponType}
    return render(request, 'coupon_management/index_couponType.html', context)

def create_couponType(request):
    if request.method == 'POST':
        form = CreateCouponTypeForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Coupon Type created successfully!')
            return redirect('couponType')
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = CreateCouponTypeForm()
    context = {'form': form}
    return render(request, 'coupon_management/create_couponType.html', context)


def view_couponType(request, coupon_type_id):
    view_couponType = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    return render(request, 'coupon_management/view_couponType.html', {'view_couponType': view_couponType})

def edit_CouponType(request, coupon_type_id):
    edit_coupon = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        form = EditCouponTypeForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('couponType')
    else:
        form = EditCouponTypeForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_couponType.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_couponType(request, coupon_type_id):
    deleteCouponType = CouponType.objects.get(coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        deleteCouponType.delete()
        return redirect('couponType')
    return render(request, 'coupon_management/delete_couponType.html', {'deleteCouponType': deleteCouponType})

#------------------------New Coupon
def new_coupon(request):
    filter_data = {}
    
    query = request.GET.get("q")
    status_type = "company"
    
    if request.GET.get('status_type'):
        status_type = request.GET.get('status_type')
    
    filter_data['status_type'] = status_type
    
    coupon_ids = CouponStock.objects.filter(coupon_stock=status_type).values_list("couponbook__pk")
    instances = NewCoupon.objects.filter(pk__in=coupon_ids).order_by("-created_date")
         
    if query:

        instances = instances.filter(
            Q(book_num__icontains=query) |
            Q(coupon_type__coupon_type_name__icontains=query)
        )
        title = "Coupon List - %s" % query
        filter_data['q'] = query
    
    context = {
        'instances': instances,
        'filter_data': filter_data
        }
    
    return render(request, 'coupon_management/index_Newcoupon.html', context)

def create_Newcoupon(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            
            coupon_type_id = request.POST.get('coupon_type')
            book_num = request.POST.get('book_num')
            valuable_leafs = request.POST.get('valuable_leafs')
            free_leafs = request.POST.get('free_leafs')
            
            selected_coupon_type = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)

            data.coupon_type = selected_coupon_type
            data.book_num = book_num
            data.no_of_leaflets = selected_coupon_type.no_of_leaflets
            data.valuable_leaflets = selected_coupon_type.valuable_leaflets
            data.free_leaflets = selected_coupon_type.free_leaflets
            data.created_by = str(request.user.id)
            
            branch_id = request.user.branch_id.branch_id
            branch = BranchMaster.objects.get(branch_id=branch_id)
            data.branch_id = branch           
            data.save()
            
            for v in valuable_leafs.split(', '):
                CouponLeaflet.objects.create(
                    coupon=data,
                    leaflet_number=data.valuable_leaflets,
                    leaflet_name=v,
                    created_by=request.user.id,
                    created_date=datetime.now(),
                )
                
            for f in free_leafs.split(', '):
                FreeLeaflet.objects.create(
                    coupon=data,
                    leaflet_number=data.free_leaflets,
                    leaflet_name=f,
                    created_by=request.user.id,
                    created_date=datetime.now(),
                )
            # Create CouponStock instance
            CouponStock.objects.create(
                couponbook=data, 
                coupon_stock='company', 
                created_by=str(request.user.id)
                )
            
            product_instance=ProdutItemMaster.objects.get(product_name=data.coupon_type.coupon_type_name)
            if (stock_intances:=ProductStock.objects.filter(product_name=product_instance,branch=branch)).exists():
                stock_intance = stock_intances.first()
                stock_intance.quantity += 1
                stock_intance.save()
            else:
                ProductStock.objects.create(
                    product_name=product_instance,
                    branch=branch,
                    quantity=1
                )

            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Coupon Generation successfully.",
                'redirect': 'true',
                "redirect_url": reverse('new_coupon')
            }
            return JsonResponse(response_data, status=200)
        else:
            message = generate_form_errors(form, formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }
            return JsonResponse(response_data, status=200)
    else:
        form = CreateNewCouponForm()
    
    context = {'form': form}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def generate_leaflets(request, coupon_id):
    coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    leaflets = []
    no_of_leaflets = int(coupon.coupon_type.no_of_leaflets)
    for leaflet_num in range(1, no_of_leaflets + 1):
        leaflet = CouponLeaflet(coupon=coupon, leaflet_number=str(leaflet_num))
        leaflets.append(leaflet)
        leaflet.save()
        

    context = {'coupon': coupon, 'leaflets': leaflets}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def get_leaflet_serial_numbers(request):
    if request.method == 'GET':
        coupon_type_id = request.GET.get('coupon_type')

        # Fetch leaflets based on the provided coupon type
        try:
            leaflets = CouponLeaflet.objects.filter(coupon__coupon_type_id=coupon_type_id)
            leaflet_data = [{'leaflet_number': leaflet.leaflet_number, 'is_used': leaflet.used} for leaflet in leaflets]
            return JsonResponse(leaflet_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

def save_coupon_data(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            try:
                # Save the coupon data to the database
                new_coupon = form.save()
                # You can also save leaflet data here if necessary
                return JsonResponse({'message': 'Coupon data saved successfully'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            return JsonResponse({'error': 'Invalid form data'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

def view_Newcoupon(request, coupon_id):
    view_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    return render(request, 'coupon_management/view_Newcoupon.html', {'view_coupon': view_coupon})

def edit_NewCoupon(request, coupon_id):
    edit_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    if request.method == 'POST':
        form = EditNewCouponForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('new_coupon')
    else:
        form = EditNewCouponForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_Newcoupon.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_Newcoupon(request, coupon_id):
    deleteCoupon = NewCoupon.objects.get(coupon_id=coupon_id)
    if request.method == 'POST':
        deleteCoupon.delete()
        return redirect('new_coupon')
    return render(request, 'coupon_management/delete_Newcoupon.html', {'deleteCoupon': deleteCoupon})


# def customer_stock(request):
#
#     couponstock=CustomerCouponStock.objects.select_related('customer').all()
#     context = {
#
#         'couponstock':couponstock
#     }
#
#     return render(request, 'coupon_management/customer_stock.html', context)


# def customer_stock(request):
#     coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock})



#
# @login_required
# def customer_stock(request):
#     # Get the user's user_type
#     user_type = request.user.user_type
#
#     # Filter customers based on user_type
#     if user_type == 'Salesman':
#         coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
#         route_li = RouteMaster.objects.filter(branch_id=request.user.branch_id)
#     else:
#         coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#         route_li = RouteMaster.objects.all()
#
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})



@login_required
def customer_stock(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()


    created_date = request.GET.get('created_date')
    if created_date:
        coupenstock = coupenstock.filter(
            coupon_type_id__created_date=created_date
        )

    return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})

@login_required
def customer_stock_coupon_details(request,customer):
    
    customer_instance = Customers.objects.get(pk=customer)
    
    customer_manual_coupons = CustomerCouponItems.objects.filter(customer_coupon__customer=customer_instance)
    
    context = {
        'customer_instance': customer_instance,
        'customer_manual_coupons': customer_manual_coupons,
    }

    return render(request, 'coupon_management/available_coupon_details.html',context
                  )



from openpyxl.styles import Font, Alignment

@login_required
def generate_excel(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_stock.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Stock'

    # Formatting styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='left', vertical='center')

    # Write headers with formatting
    headers = [
        'Sl.No', 'Customer Name / Mobile No', 'Building Name / House No',
        'Digital Coupon Count', 'Manual Coupon Count', 'Total Count'
    ]
    for col_num, header_title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data with formatting
    for row_num, stock in enumerate(coupenstock, start=2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1).alignment = data_alignment
        worksheet.cell(row=row_num, column=2, value=stock.customer.customer_name + ', ' + stock.customer.mobile_no).alignment = data_alignment
        worksheet.cell(row=row_num, column=3, value=stock.customer.building_name + ', ' + stock.customer.door_house_no).alignment = data_alignment
        if stock.coupon_method == 'digital':
            worksheet.cell(row=row_num, column=4, value=stock.count).alignment = data_alignment
        else:
            worksheet.cell(row=row_num, column=5, value=stock.count).alignment = data_alignment

    # Calculate and write total counts
    total_digital_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'digital')
    total_manual_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'manual')
    total_row_num = len(coupenstock) + 2
    worksheet.cell(row=total_row_num, column=4, value=total_digital_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=5, value=total_manual_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=6, value=total_digital_count + total_manual_count).alignment = data_alignment

    # Autofit column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response

from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


def customer_stock_pdf(request):
    # Retrieve customer stock data
    coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    # Check if coupenstock is not empty
    if coupenstock:
        # Create the HTTP response with PDF content type and attachment filename
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customer_stock.pdf"'

        # Create a PDF document
        pdf_buffer = BytesIO()
        pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        data = []

        # Add headers
        headers = ['Sl.No', 'Customer Name / Mobile No', 'Building Name / House No', 'Digital', 'Manual', 'Total Count']
        data.append(headers)

        # Add data to the PDF document
        for index, stock in enumerate(coupenstock):
            customer_name_mobile = f"{stock.customer.customer_name}, {stock.customer.mobile_no}"
            building_house = f"{stock.customer.building_name}, {stock.customer.door_house_no}"
            digital = stock.count if stock.coupon_method == 'digital' else ''
            manual = stock.count if stock.coupon_method == 'manual' else ''
            total_count = stock.count
            data.append([index + 1, customer_name_mobile, building_house, digital, manual, total_count])

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf_value = pdf_buffer.getvalue()
        pdf_buffer.close()
        response.write(pdf_value)

        return response
    else:
        # Return an empty HTTP response with a message indicating no data available
        return HttpResponse('No customer stock data available.')
    


def redeemed_history(request):
    filter_data = {}
    query = request.GET.get("q", "").strip()  # Get the search query

    start_date = date.today()
    end_date = date.today()
    route_name =request.GET.get('route_name')
    
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    filter_data["start_date"] = start_date.strftime('%Y-%m-%d')
    filter_data["end_date"] = end_date.strftime('%Y-%m-%d')
    
    instances = CustomerSupply.objects.filter(created_date__date__gte=start_date,created_date__date__lte=end_date,customer__sales_type="CASH COUPON").order_by("-created_date")
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
    # Apply search filter to instances if a query exists
    if query:
        instances = instances.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__location__location_name__icontains=query) |
            Q(customer__building_name__icontains=query)
        )
        filter_data['q'] = query
        
    # Calculate totals for manual and digital coupons manually
    total_manual_coupons = 0
    total_digital_coupons = 0
    for instance in instances:
        total_coupons = instance.total_coupon_recieved()  # Using the method from the model
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)
        
    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()
    context = {
        'instances': instances,
        "filter_data": filter_data,
        "route_li": route_li,
        "total_manual_coupons": total_manual_coupons,
        "total_digital_coupons": total_digital_coupons,
    }

    return render(request, 'coupon_management/redeemed_history.html', context)

def print_redeemed_history(request):
    filter_data = {}
    query = request.GET.get("q", "").strip()  # Get the search query

    start_date = date.today()
    end_date = date.today()
    route_name =request.GET.get('route_name')
    
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    filter_data["start_date"] = start_date.strftime('%Y-%m-%d')
    filter_data["end_date"] = end_date.strftime('%Y-%m-%d')
    
    instances = CustomerSupply.objects.filter(created_date__date__gte=start_date,created_date__date__lte=end_date,customer__sales_type="CASH COUPON").order_by("-created_date")
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
    # Apply search filter to instances if a query exists
    if query:
        instances = instances.filter(
            Q(customer__custom_id__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__location__location_name__icontains=query) |
            Q(customer__building_name__icontains=query)
        )
        filter_data['q'] = query
        
    # Calculate totals for manual and digital coupons manually
    total_manual_coupons = 0
    total_digital_coupons = 0
    for instance in instances:
        total_coupons = instance.total_coupon_recieved()  # Using the method from the model
        total_manual_coupons += total_coupons.get('manual_coupon', 0)
        total_digital_coupons += total_coupons.get('digital_coupon', 0)
        
    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()
    context = {
        'instances': instances,
        "filter_data": filter_data,
        "route_li": route_li,
        "total_manual_coupons": total_manual_coupons,
        "total_digital_coupons": total_digital_coupons,
    }

    return render(request, 'coupon_management/print_redeemed_history.html', context)
